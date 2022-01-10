""" This script will extract the customer service offering information that is stored in SNOW
"""
from openpyxl import load_workbook


def get_worksheet_table(worksheet, table_name):
    """ Get all tables from a given workbook. Returns a dictionary of tables.
        Requires a filename, which includes the file path and filename. """

    # Load the workbook, from the filename
    # wb = load_workbook(filename=filename, read_only=False, keep_vba=False, data_only=True, keep_links=False)

    # Initialize the dictionary of tables
    service_offerings_list = []

    for tbl in worksheet._tables:
        if tbl.name == table_name:
            # Grab the 'data' from the table
            data = worksheet[tbl.ref]

            # Now convert the table 'data' to a Pandas DataFrame
            # First get a list of all rows, including the first header row
            rows_list = []
            for row in data:
                # Get a list of all columns in each row
                cols = []
                for col in row:
                    cols.append(col.value)
                rows_list.append(cols)
    # for column_name in rows_list[0]:
    #     print("Column name:" + str(column_name))
    # reader = csv.DictReader((row[0] for row in rows_list))
    # for offering_row in reader:
    #     print("Number: " + str(offering_row['Number']) + "Name: " + str(offering_row['Name']))
    header_columns = [col.upper() for col in rows_list[0]]
    for data_row in rows_list[1:]:
        service_offering_dict = zip(header_columns, data_row)
        service_offerings_list.append(dict(service_offering_dict))

    return service_offerings_list


def list_defined_tables(worksheet):
    for tbl in worksheet._tables:
        print(" : " + tbl.displayName)
        print("   -  name = " + tbl.name)
        print("   -  type = " + tbl.tableType if isinstance(tbl.tableType, str) else 'n/a')
        print("   - range = " + tbl.ref)
        print("   - #cols = %d" % len(tbl.tableColumns))
        for col in tbl.tableColumns:
            print("     : " + col.name)


if __name__ == "__main__":
    service_offering_excel_workbook = "C:\\Users\\dhartman\\Documents\\Support\\AWS-Resources-Production.xlsx"
    # Run the function to return a dictionary of all tables in the Excel workbook
    # tables_dict = get_all_tables(filename=service_offering_excel_workbook)

    wb = load_workbook(filename=service_offering_excel_workbook, data_only=True)
    ws = wb["ServiceOfferings"]
    service_offerings_table = get_worksheet_table(ws, "SERVICE_OFFERINGS")
    filtered_offerings_list = [row for row in service_offerings_table
                               if "INNIO" in row['COMPANY'].upper()]
    print(str(filtered_offerings_list))
